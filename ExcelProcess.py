import openpyxl
import numpy as np
from openpyxl import Workbook
import os


def __readExcel__(path, sheetName=''):
    wb = openpyxl.load_workbook(r'' + path, read_only=True)
    # 如果没传入工作簿名称，使用默认工作簿
    if sheetName.__eq__(""):
        return wb.active
    else:
        return wb.get_sheet_by_name(sheetName)


# 返回表格所有的工作簿名称
def getAllSheet(path):
    wb = openpyxl.load_workbook(r'' + path, read_only=True)
    return wb.sheetnames


# 返回表格的内容，格式为双重列表 如果sheetName为all，则把所有工作簿都返回，格式为字典{"工作簿名":数据}
def excelToList(path, sheetName=''):
    # 返回所有工作簿的内容
    if sheetName.__eq__("all"):
        di = {}
        for i in getAllSheet(path):
            di[i] = excelToList(path, i)
        return di
    li = []
    # 打开文件
    ws = __readExcel__(path, sheetName)
    for i in ws.values:
        li.append(list(i))
    return li


# 把列表转换为表格形式，如果是一维列表，则插入一行数据。如果是二维列表，内层列表为一行数据。
# 如果文件已存在，则复制源文件，然后添加新文件
def listToExcel(data, path, sheetName=""):
    # 如果列表为空
    if len(data) == 0:
        return "列表为空"
    wb = Workbook()
    print(wb.sheetnames)
    # 如果文件已经存在，则则复制源文件，然后添加新文件
    if os.path.exists(path):
        # 获得所有工作簿内容，为字典
        allData = excelToList(path, 'all')
        for i in allData:
            if i not in wb.sheetnames:
                wb.create_sheet(i)
            ws = wb.get_sheet_by_name(i)
            for k in allData[i]:
                ws.append(k)
    # 如果没输入工作簿名称，使用默认
    if sheetName.__eq__(""):
        ws = wb.active
    else:
        if sheetName not in wb.sheetnames:
            wb.create_sheet(sheetName)
        ws = wb.get_sheet_by_name(sheetName)
    # 判断列表的维度，不支持三维以上，因为表格本身就是二维
    if len([i for i in str(np.array(data).shape)[:-1][1:].replace(" ", "").split(",") if not i.__eq__("")]) == 1:
        # 一维
        ws.append(data)
    elif len([i for i in str(np.array(data).shape)[:-1][1:].replace(" ", "").split(",") if not i.__eq__("")]) == 2:
        # 二维
        for i in data:
            ws.append(list(i))
    else:
        # 大于二维
        return "只支持一维、二维列表"
    wb.save(path)
    wb.close()
